"""
Generate DB Tables Report Excel — PetCircle Condition Architecture
Shows exactly what rows would be written to each table.
Outputs: D:/pet-circle/backend/DB_Tables_Report.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ─── Colours ─────────────────────────────────────────────────────────────────
ORANGE       = "D44800"
DARK_GREY    = "2D2D2D"
MID_GREY     = "F2F2F2"
LIGHT_ORANGE = "FFF0E8"
WHITE        = "FFFFFF"
GREEN        = "E8F5E9"
BLUE         = "E3F2FD"
YELLOW       = "FFF9C4"
HEADER_GREY  = "424242"

def _fill(c): return PatternFill("solid", fgColor=c)
def _font(bold=False, color="000000", size=9): return Font(bold=bold, color=color, size=size, name="Calibri")
def _border():
    s = Side(style="thin", color="BBBBBB")
    return Border(left=s, right=s, top=s, bottom=s)

def _write(ws, r, c, val, bg=WHITE, bold=False, color="000000", size=9, wrap=True, align="left"):
    cell = ws.cell(row=r, column=c, value=val)
    cell.fill = _fill(bg)
    cell.font = _font(bold=bold, color=color, size=size)
    cell.alignment = Alignment(horizontal=align, vertical="top", wrap_text=wrap)
    cell.border = _border()
    return cell

def _title_block(ws, table_name, subtitle, start_row=1):
    ws.merge_cells(f"A{start_row}:{get_column_letter(20)}{start_row}")
    c = ws.cell(row=start_row, column=1, value=f"  TABLE: {table_name}")
    c.fill = _fill(ORANGE); c.font = Font(bold=True, color=WHITE, size=13, name="Calibri")
    c.alignment = Alignment(vertical="center"); ws.row_dimensions[start_row].height = 28

    ws.merge_cells(f"A{start_row+1}:{get_column_letter(20)}{start_row+1}")
    c2 = ws.cell(row=start_row+1, column=1, value=f"  {subtitle}")
    c2.fill = _fill(DARK_GREY); c2.font = Font(color=WHITE, size=9, name="Calibri")
    c2.alignment = Alignment(vertical="center"); ws.row_dimensions[start_row+1].height = 16
    return start_row + 2

def _col_headers(ws, row, headers, bg=HEADER_GREY):
    for col, h in enumerate(headers, 1):
        _write(ws, row, col, h, bg=bg, bold=True, color=WHITE, size=9, align="center")
    ws.row_dimensions[row].height = 22
    return row + 1

def _note(ws, row, ncols, text, bg=MID_GREY):
    ws.merge_cells(f"A{row}:{get_column_letter(ncols)}{row+2}")
    c = ws.cell(row=row, column=1, value=text)
    c.fill = _fill(bg); c.font = _font(size=8, color="444444")
    c.alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[row].height = 50

def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ─── Shortened IDs for readability ────────────────────────────────────────────
# Real UUIDs would be gen_random_uuid() — we use short labels to show structure
PET_ID   = "pet-0001-xxxx"
DOC_A    = "doc-4cc6-xxxx"   # 4cc61ecc.JPG  (Ear mites / Pyoderma / Fungal)
DOC_B    = "doc-887a-xxxx"   # 887afa02.JPG  (Post-neuter care)
DOC_C    = "doc-f4d4-xxxx"   # f4d47d0c.JPG  (Pyoderma / Fungal)

COND_1   = "cond-0001"       # Ear mites         (doc A)
COND_2   = "cond-0002"       # Pyoderma          (doc A)
COND_3   = "cond-0003"       # Fungal infection  (doc A)
COND_4   = "cond-0004"       # Post-neuter care  (doc B)
COND_5   = "cond-0005"       # Pyoderma          (doc C)
COND_6   = "cond-0006"       # Fungal infection  (doc C)

AGG_1    = "agg-fam-0001"    # Ear mites family
AGG_2    = "agg-fam-0002"    # Pyoderma/Fungal family
AGG_3    = "agg-fam-0003"    # Post-neuter care family

MED_1    = "med-0001"
MED_2    = "med-0002"
MED_3    = "med-0003"
MED_4    = "med-0004"
MED_5    = "med-0005"
MED_6    = "med-0006"
MED_7    = "med-0007"
MED_8    = "med-0008"
MED_9    = "med-0009"

MON_1    = "mon-0001"
MON_2    = "mon-0002"
MON_3    = "mon-0003"


wb = Workbook()
wb.remove(wb.active)


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 1 — conditions
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("conditions")
r = _title_block(ws, "conditions",
    "One row per named condition per uploaded document. Medications & monitoring stored in their own tables (linked by condition_id).")

headers = ["id", "pet_id", "document_id", "name", "condition_type",
           "condition_status", "episode_dates (JSONB)", "diagnosed_at",
           "is_active", "condition_family_id", "recurrence_watch",
           "source", "notes / managed_by"]
r = _col_headers(ws, r, headers)

rows = [
    (COND_1, PET_ID, DOC_A, "Ear mites",        "episodic", "active",  "[]",              "NULL",       "true",  AGG_1, "false", "extraction", "NULL"),
    (COND_2, PET_ID, DOC_A, "Pyoderma",          "episodic", "active",  "[]",              "NULL",       "true",  AGG_2, "false", "extraction", "NULL"),
    (COND_3, PET_ID, DOC_A, "Fungal infection",  "episodic", "active",  "[]",              "NULL",       "true",  AGG_2, "false", "extraction", "NULL"),
    (COND_4, PET_ID, DOC_B, "Post-neuter care",  "episodic", "active",  '["31/1/23"]',     "2023-01-31", "true",  AGG_3, "false", "extraction", "NULL"),
    (COND_5, PET_ID, DOC_C, "Pyoderma",          "episodic", "active",  '["2024-12-19"]',  "2024-12-19", "true",  AGG_2, "false", "extraction", "NULL"),
    (COND_6, PET_ID, DOC_C, "Fungal infection",  "episodic", "active",  '["2024-12-19"]',  "2024-12-19", "true",  AGG_2, "false", "extraction", "NULL"),
]
stripe = [WHITE, LIGHT_ORANGE]
for i, row in enumerate(rows):
    bg = stripe[i % 2]
    for col, val in enumerate(row, 1):
        _write(ws, r, col, val, bg=bg)
    ws.row_dimensions[r].height = 18
    r += 1

r += 1
_note(ws, r, len(headers),
    "NOTES:\n"
    "• condition_type is ONLY 'chronic' or 'episodic' at document level — 'recurrent' is never stored here\n"
    "• condition_status shown as 'active' (written at extraction time) but this is NEVER read by the dashboard — it is recomputed at runtime\n"
    "• condition_family_id is written back AFTER aggregation runs (NULL until aggregation completes)\n"
    "• Pyoderma + Fungal infection from doc A have no episode_dates because that document had no date text")

_set_col_widths(ws, [14, 14, 14, 20, 14, 14, 20, 13, 9, 14, 14, 12, 20])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 2 — condition_medications
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("condition_medications")
r = _title_block(ws, "condition_medications",
    "One row per drug per condition row. Linked to conditions via condition_id FK.")

headers = ["id", "condition_id", "condition name\n(for reference)", "name (drug)",
           "item_type", "dose", "frequency", "route", "end_date", "started_at",
           "status (legacy)", "refill_due_date", "price", "notes"]
r = _col_headers(ws, r, headers)

rows = [
    # Post-neuter care (cond-0004)
    (MED_1, COND_4, "Post-neuter care", "Riovet spray",    "medicine", "NULL",  "NULL", "topical", "NULL",       "NULL",       "active", "NULL", "NULL", "NULL"),
    (MED_2, COND_4, "Post-neuter care", "WND Heal powder", "medicine", "NULL",  "NULL", "topical", "NULL",       "NULL",       "active", "NULL", "NULL", "NULL"),
    (MED_3, COND_4, "Post-neuter care", "Caprill 100",     "medicine", "0-1",   "NULL", "NULL",    "2023-02-10", "2023-01-31", "active", "NULL", "NULL", "NULL"),
    (MED_4, COND_4, "Post-neuter care", "Cefpet CV",       "medicine", "0-1",   "NULL", "NULL",    "2023-02-10", "2023-01-31", "active", "NULL", "NULL", "NULL"),
    (MED_5, COND_4, "Post-neuter care", "Pantopet 40",     "medicine", "0-1",   "NULL", "NULL",    "2023-02-10", "2023-01-31", "active", "NULL", "NULL", "NULL"),
    # Pyoderma doc C (cond-0005)
    (MED_6, COND_5, "Pyoderma (doc C)", "Itralet",         "medicine", "NULL",  "NULL", "oral",    "2024-12-29", "2024-12-19", "active", "NULL", "NULL", "NULL"),
    (MED_7, COND_5, "Pyoderma (doc C)", "Panbact D",       "medicine", "NULL",  "NULL", "NULL",    "2024-12-29", "2024-12-19", "active", "NULL", "NULL", "NULL"),
    (MED_8, COND_5, "Pyoderma (doc C)", "Onfyte lotion",   "medicine", "NULL",  "NULL", "topical", "2024-12-29", "2024-12-19", "active", "NULL", "NULL", "NULL"),
    # Fungal infection doc C (cond-0006)
    (MED_9, COND_6, "Fungal inf. (doc C)", "Onfyte lotion","medicine", "NULL",  "NULL", "topical", "2024-12-29", "2024-12-19", "active", "NULL", "NULL", "NULL"),
    # Ear mites / Pyoderma doc A / Fungal doc A → NO medications extracted from that document
]
group_colours = [LIGHT_ORANGE, LIGHT_ORANGE, LIGHT_ORANGE, LIGHT_ORANGE, LIGHT_ORANGE,
                 BLUE, BLUE, BLUE, GREEN]
for i, (row, bg) in enumerate(zip(rows, group_colours)):
    for col, val in enumerate(row, 1):
        _write(ws, r, col, val, bg=bg)
    ws.row_dimensions[r].height = 18
    r += 1

r += 1
_note(ws, r, len(headers),
    "NOTES:\n"
    "• Ear mites (cond-0001), Pyoderma doc A (cond-0002), Fungal doc A (cond-0003) → 0 medication rows (no medications in that document)\n"
    "• 'status' column defaults to 'active' and is NEVER updated — active status is computed from end_date at runtime via is_medication_active()\n"
    "• Onfyte lotion appears in BOTH cond-0005 and cond-0006 — these are two separate rows; deduplication happens in aggregated_conditions.medications JSONB")

_set_col_widths(ws, [10, 12, 20, 18, 10, 8, 8, 9, 13, 13, 12, 13, 8, 12])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 3 — condition_monitoring
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("condition_monitoring")
r = _title_block(ws, "condition_monitoring",
    "One row per follow-up check per condition row.")

headers = ["id", "condition_id", "condition name\n(for reference)", "name",
           "frequency", "recheck_due_date", "next_due_date", "last_done_date", "result_summary"]
r = _col_headers(ws, r, headers)

rows = [
    (MON_1, COND_4, "Post-neuter care", "Recheck", "NULL", "2023-02-08", "NULL", "NULL", "NULL"),
    (MON_2, COND_4, "Post-neuter care", "Recheck", "NULL", "2023-02-13", "NULL", "NULL", "NULL"),
    (MON_3, COND_5, "Pyoderma (doc C)", "Recheck", "NULL", "2024-12-29", "NULL", "NULL", "NULL"),
]
stripe = [LIGHT_ORANGE, LIGHT_ORANGE, BLUE]
for i, (row, bg) in enumerate(zip(rows, stripe)):
    for col, val in enumerate(row, 1):
        _write(ws, r, col, val, bg=bg)
    ws.row_dimensions[r].height = 18
    r += 1

r += 1
_note(ws, r, len(headers),
    "NOTES:\n"
    "• Two Recheck rows for Post-neuter care because the document mentioned two follow-up dates (8/2/23 and 13/2/23)\n"
    "• Ear mites, Pyoderma doc A, Fungal infection docs A & C → 0 monitoring rows")

_set_col_widths(ws, [10, 12, 22, 14, 12, 16, 14, 14, 18])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 4 — aggregated_conditions
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("aggregated_conditions")
r = _title_block(ws, "aggregated_conditions",
    "One row per complaint family per pet. Medications & monitoring stored inline as JSONB (merged/deduplicated from all docs in the family).")

headers = ["id", "pet_id", "name", "condition_type\n(family level)",
           "condition_status\n(stored — NOT used at runtime)",
           "episode_dates\n(JSONB merged)", "diagnosed_at\n(MIN)",
           "last_record_date\n(MAX episode)", "medication_end_date\n(MAX across family)",
           "medications\n(JSONB — merged)", "monitoring\n(JSONB — merged)",
           "soft_resolution", "recurrence_watch", "canonical_condition_id\n(→ conditions)"]
r = _col_headers(ws, r, headers)

rows = [
    (
        AGG_1, PET_ID,
        "Ear mites",
        "episodic",
        "resolved",
        "[]",
        "NULL",
        "NULL",
        "NULL",
        "[]",
        "[]",
        "false",
        "true",
        COND_1,
    ),
    (
        AGG_2, PET_ID,
        "Pyoderma",
        "episodic",
        "resolved",
        '["2024-12-19"]',
        "2024-12-19",
        "2024-12-19",
        "2024-12-29",
        '[{"name":"Itralet","end_date":"2024-12-29","dose":null,"frequency":null},'
        '{"name":"Panbact D","end_date":"2024-12-29","dose":null,"frequency":null},'
        '{"name":"Onfyte lotion","end_date":"2024-12-29","dose":null,"frequency":null}]',
        '[{"name":"Recheck","recheck_due_date":"2024-12-29"}]',
        "false",
        "true",
        COND_5,
    ),
    (
        AGG_3, PET_ID,
        "Post-neuter care",
        "episodic",
        "resolved",
        '["2023-01-31"]',
        "2023-01-31",
        "2023-01-31",
        "2023-02-10",
        '[{"name":"Riovet spray","end_date":null,"dose":null,"frequency":null},'
        '{"name":"WND Heal powder","end_date":null,"dose":null,"frequency":null},'
        '{"name":"Caprill 100","end_date":"2023-02-10","dose":"0-1","frequency":null},'
        '{"name":"Cefpet CV","end_date":"2023-02-10","dose":"0-1","frequency":null},'
        '{"name":"Pantopet 40","end_date":"2023-02-10","dose":"0-1","frequency":null}]',
        '[{"name":"Recheck","recheck_due_date":"2023-02-13"}]',
        "false",
        "true",
        COND_4,
    ),
]
colours = [LIGHT_ORANGE, BLUE, GREEN]
for row, bg in zip(rows, colours):
    for col, val in enumerate(row, 1):
        _write(ws, r, col, val, bg=bg)
    ws.row_dimensions[r].height = 80
    r += 1

r += 1
_note(ws, r, len(headers),
    "NOTES:\n"
    "• AGG_2 (Pyoderma family) contains 4 document-level rows: Pyoderma(docA) + Fungal(docA) + Pyoderma(docC) + Fungal(docC) — same body system (skin) + same sub-pattern (infection)\n"
    "• Onfyte lotion deduplicated (appeared in cond-0005 AND cond-0006) — MAX end_date kept → 2024-12-29\n"
    "• monitoring Recheck: two recheck dates in Post-neuter care (8/2/23, 13/2/23) — MAX kept → 2023-02-13\n"
    "• condition_status stored as 'resolved' but dashboard NEVER reads this column — it calls _compute_condition_status(type, medication_end_date, episode_dates) fresh on every load\n"
    "• canonical_condition_id → the doc-level row with MIN diagnosed_at in this family")

_set_col_widths(ws, [13, 13, 18, 14, 16, 18, 13, 15, 18, 55, 45, 13, 14, 14])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 5 — diagnostic_test_results (lab)
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("diagnostic_test_results (lab)")
r = _title_block(ws, "diagnostic_test_results",
    "Lab / blood / urine test values. Completely separate from the conditions pipeline — linked only by document_id.")

headers = ["id", "pet_id", "document_id", "test_type", "parameter_name",
           "value_numeric", "value_text", "unit", "reference_range",
           "status_flag", "observed_at"]
r = _col_headers(ws, r, headers)

# Sample rows (these are illustrative — none of the 10 condition docs happened to be blood reports,
# but this is exactly how they'd look if they were)
sample_rows = [
    ("lab-0001", PET_ID, "doc-blood-xxxx", "blood", "Haemoglobin",  "12.4",  "12.4",    "g/dL",     "12.0 – 18.0", "normal",   "2024-11-15"),
    ("lab-0002", PET_ID, "doc-blood-xxxx", "blood", "WBC",           "8.2",   "8.2",     "×10³/μL",  "6.0 – 17.0",  "normal",   "2024-11-15"),
    ("lab-0003", PET_ID, "doc-blood-xxxx", "blood", "Platelets",     "180",   "180",     "×10³/μL",  "170 – 400",   "normal",   "2024-11-15"),
    ("lab-0004", PET_ID, "doc-blood-xxxx", "blood", "Creatinine",    "1.8",   "1.8",     "mg/dL",    "0.5 – 1.5",   "high",     "2024-11-15"),
    ("lab-0005", PET_ID, "doc-blood-xxxx", "blood", "BUN",           "32.0",  "32.0",    "mg/dL",    "7 – 25",      "high",     "2024-11-15"),
    ("lab-0006", PET_ID, "doc-urine-xxxx", "urine", "Protein",       "NULL",  "Positive","NULL",      "Negative",    "abnormal", "2024-11-15"),
    ("lab-0007", PET_ID, "doc-urine-xxxx", "urine", "Glucose",       "NULL",  "Negative","NULL",      "Negative",    "normal",   "2024-11-15"),
    ("lab-0008", PET_ID, "doc-vital-xxxx", "vital", "Weight",        "8.5",   "8.5 kg",  "kg",        "NULL",        "NULL",     "2024-11-15"),
]
stripe = [WHITE, LIGHT_ORANGE]
for i, row in enumerate(sample_rows):
    bg = stripe[i % 2]
    for col, val in enumerate(row, 1):
        _write(ws, r, col, val, bg=bg)
    ws.row_dimensions[r].height = 16
    r += 1

r += 1
_note(ws, r, len(headers),
    "NOTES:\n"
    "• These are ILLUSTRATIVE rows — none of the 10 condition test documents were blood/urine reports, so no lab rows were extracted in this test run\n"
    "• Lab results are COMPLETELY SEPARATE from conditions — no FK to conditions table; both tables share document_id → documents.id\n"
    "• value_numeric = NULL when result is text-only (e.g. Positive/Negative); value_text always has the raw string\n"
    "• status_flag (low/normal/high/abnormal) is derived by GPT comparing value to reference_range")

_set_col_widths(ws, [10, 13, 15, 10, 22, 14, 14, 12, 16, 12, 13])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 6 — FK relationship map
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("FK Relationships")
r = _title_block(ws, "Foreign Key Relationships",
    "How all five tables connect — read top to bottom for data flow")

headers = ["From Table", "Column", "→", "To Table", "Column", "On Delete", "Direction / Purpose"]
r = _col_headers(ws, r, headers)

fk_rows = [
    ("conditions",           "pet_id",                "→", "pets",                 "id",  "CASCADE",  "Every condition belongs to one pet"),
    ("conditions",           "document_id",           "→", "documents",            "id",  "SET NULL", "Source document (extraction origin); nullable"),
    ("conditions",           "condition_family_id",   "→", "aggregated_conditions","id",  "SET NULL", "Written back by aggregation service after grouping"),
    ("condition_medications","condition_id",           "→", "conditions",           "id",  "CASCADE",  "Medications belong to one document-level condition row"),
    ("condition_monitoring", "condition_id",           "→", "conditions",           "id",  "CASCADE",  "Monitoring items belong to one document-level condition row"),
    ("aggregated_conditions","pet_id",                "→", "pets",                 "id",  "CASCADE",  "Each aggregated family belongs to one pet"),
    ("aggregated_conditions","canonical_condition_id","→", "conditions",           "id",  "SET NULL", "Points to earliest-diagnosed doc-level row in the family"),
    ("diagnostic_test_results","pet_id",             "→", "pets",                 "id",  "CASCADE",  "Lab result belongs to one pet"),
    ("diagnostic_test_results","document_id",         "→", "documents",            "id",  "SET NULL", "Source report document; no FK to conditions"),
]
stripe = [WHITE, LIGHT_ORANGE]
for i, row in enumerate(fk_rows):
    bg = stripe[i % 2]
    for col, val in enumerate(row, 1):
        _write(ws, r, col, val, bg=bg)
    ws.row_dimensions[r].height = 18
    r += 1

_set_col_widths(ws, [26, 26, 4, 26, 12, 12, 45])


# ══════════════════════════════════════════════════════════════════════════════
# SHEET 7 — condition_status runtime logic
# ══════════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("condition_status Logic")
r = _title_block(ws, "condition_status — Runtime Computation",
    "Never read from DB. Computed on every dashboard load by _compute_condition_status(type, medication_end_date, episode_dates)")

headers = ["condition_type", "medication_end_date", "episode_dates", "Rule", "→ condition_status"]
r = _col_headers(ws, r, headers)

logic_rows = [
    ("chronic",              "any",          "any",                    "chronic is always active",                                          "active"),
    ("episodic / recurrent", ">= today",     "any",                    "medication still running",                                          "active"),
    ("episodic / recurrent", "< today",      "any",                    "med ended ≤ 30 days ago",                                           "monitoring"),
    ("episodic / recurrent", "NULL",         "latest ≤ 30 days ago",   "no end_date, recent episode",                                       "active"),
    ("episodic / recurrent", "NULL",         "30 < latest ≤ 60 days",  "no end_date, recent-ish episode",                                   "monitoring"),
    ("episodic / recurrent", "NULL or old",  "latest > 60 days ago",   "all evidence is old",                                               "resolved"),
]
status_bg = {"active": GREEN, "monitoring": YELLOW, "resolved": MID_GREY}
for row in logic_rows:
    bg = status_bg.get(row[-1], WHITE)
    for col, val in enumerate(row, 1):
        _write(ws, r, col, val, bg=bg)
    ws.row_dimensions[r].height = 20
    r += 1

r += 2
ws.merge_cells(f"A{r}:E{r+4}")
c = ws.cell(row=r, column=1, value=(
    "RECURRENCE UPGRADE (applied at family level during aggregation — upgrades condition_type in aggregated_conditions)\n\n"
    "≥ 2 unique episode dates within last 15 months  →  condition_type = 'recurrent'\n"
    "≥ 3 unique episode dates within last 24 months  →  condition_type = 'recurrent'\n"
    "1 unique episode only                           →  condition_type = 'episodic', recurrence_watch = true\n\n"
    "'recurrent' is NEVER stored in the conditions table (document level). It only appears in aggregated_conditions."
))
c.fill = _fill(MID_GREY); c.font = _font(size=9); c.alignment = Alignment(wrap_text=True, vertical="top")
ws.row_dimensions[r].height = 90

_set_col_widths(ws, [22, 22, 26, 40, 20])


# ── Save ──────────────────────────────────────────────────────────────────────
out_path = r"D:\pet-circle\backend\DB_Tables_Report.xlsx"
wb.save(out_path)
print(f"Saved: {out_path}")
