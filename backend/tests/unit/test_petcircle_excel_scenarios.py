from __future__ import annotations

import re
from functools import lru_cache
from pathlib import Path

import pytest
from openpyxl import load_workbook

REPO_ROOT = Path(__file__).resolve().parents[3]
TEST_CASES_XLSX = REPO_ROOT / "project details" / "petcircle_test_cases.xlsx"
ALL_TEST_CASES_SHEET = "All Test Cases"
SUMMARY_DASHBOARD_SHEET = "Summary Dashboard"
CONVERSATION_SCRIPTS_SHEET = "Conversation Scripts"
LEGEND_NOTES_SHEET = "Legend & Notes"

REQUIRED_SHEETS = (
    ALL_TEST_CASES_SHEET,
    SUMMARY_DASHBOARD_SHEET,
    CONVERSATION_SCRIPTS_SHEET,
    LEGEND_NOTES_SHEET,
)

REQUIRED_COLUMNS = (
    "Test ID",
    "Category",
    "Test Title",
    "Description",
    "Simulated User Input",
    "Expected System Behaviour",
    "Priority",
    "Flow Step",
)

EXPECTED_CATEGORY_BY_PREFIX = {
    "IF": "Input format",
    "EC": "Edge / ambiguous",
    "NU": "Nutrition gap",
    "UI": "User interruption",
    "TD": "Tech disruption",
    "DO": "Document upload",
}

VALID_PRIORITIES = {"High", "Medium", "Low"}

SCENARIO_ID_PATTERN = re.compile(r"^(IF|EC|NU|UI|TD|DO)-(\d{2})$")
SCENARIO_ID_IN_TEXT_PATTERN = re.compile(r"\b(IF|EC|NU|UI|TD|DO)-(\d{2})\b")

EXPECTED_TEST_IDS = {
    "IF-01", "IF-02", "IF-03", "IF-04", "IF-05", "IF-06", "IF-07", "IF-08", "IF-09", "IF-10", "IF-11", "IF-12",
    "EC-01", "EC-02", "EC-03", "EC-04", "EC-05", "EC-06", "EC-07", "EC-08", "EC-09", "EC-10",
    "NU-01", "NU-02", "NU-03", "NU-04", "NU-05",
    "UI-01", "UI-02", "UI-03", "UI-04", "UI-05", "UI-06", "UI-07", "UI-08",
    "TD-01", "TD-02", "TD-03", "TD-04", "TD-05",
    "DO-01", "DO-02", "DO-03", "DO-04", "DO-05", "DO-06",
}


@lru_cache(maxsize=1)
def _load_case_rows() -> list[dict[str, str]]:
    """Load all non-empty scenario rows from the Excel test case sheet."""
    workbook = load_workbook(TEST_CASES_XLSX)
    try:
        sheet = workbook[ALL_TEST_CASES_SHEET]

        headers = [cell.value for cell in sheet[1]]
        rows: list[dict[str, str]] = []

        for row in sheet.iter_rows(min_row=2, values_only=True):
            if not any(value is not None and str(value).strip() for value in row):
                continue
            rows.append({headers[i]: row[i] for i in range(len(headers))})

        return rows
    finally:
        workbook.close()


@lru_cache(maxsize=1)
def _load_conversation_script_case_ids() -> list[str]:
    """Extract scenario IDs referenced in the conversation script examples."""
    workbook = load_workbook(TEST_CASES_XLSX)
    try:
        sheet = workbook[CONVERSATION_SCRIPTS_SHEET]
        case_ids: set[str] = set()

        for row in sheet.iter_rows(min_row=1, values_only=True):
            for cell in row:
                if cell is None:
                    continue
                text = str(cell).strip()
                for match in SCENARIO_ID_IN_TEXT_PATTERN.findall(text):
                    case_ids.add(f"{match[0]}-{match[1]}")

        return sorted(case_ids)
    finally:
        workbook.close()


def _non_empty(value: object) -> bool:
    return value is not None and bool(str(value).strip())


def _to_int(value: object) -> int | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        as_float = float(value)
        return int(as_float) if as_float.is_integer() else None
    text = str(value).strip()
    if not text:
        return None
    try:
        as_float = float(text)
        return int(as_float) if as_float.is_integer() else None
    except ValueError:
        return None


def test_excel_file_and_sheet_exist() -> None:
    assert TEST_CASES_XLSX.exists(), f"Missing test workbook: {TEST_CASES_XLSX}"

    workbook = load_workbook(TEST_CASES_XLSX)
    try:
        assert set(REQUIRED_SHEETS).issubset(set(workbook.sheetnames))
    finally:
        workbook.close()


def test_all_expected_scenarios_present_and_unique() -> None:
    cases = _load_case_rows()
    found_ids = [str(case.get("Test ID") or "").strip() for case in cases if _non_empty(case.get("Test ID"))]
    assert len(found_ids) == len(set(found_ids)), "Duplicate Test ID values found in workbook"
    assert set(found_ids) == EXPECTED_TEST_IDS


def test_summary_dashboard_category_and_total_counts_match_all_test_cases() -> None:
    cases = _load_case_rows()
    all_case_categories = [
        str(case.get("Category") or "").strip()
        for case in cases
        if _non_empty(case.get("Test ID")) and _non_empty(case.get("Category"))
    ]

    expected_counts = {
        category: all_case_categories.count(category)
        for category in EXPECTED_CATEGORY_BY_PREFIX.values()
    }
    expected_total = len(all_case_categories)

    workbook = load_workbook(TEST_CASES_XLSX)
    try:
        sheet = workbook[SUMMARY_DASHBOARD_SHEET]
        observed_counts: dict[str, int] = {}
        observed_total: int | None = None
        total_uses_formula = False
        total_formula_text: str | None = None
        unknown_labels: list[str] = []

        in_category_table = False
        for row in sheet.iter_rows(min_row=1, values_only=True):
            label = str(row[0]).strip() if row and row[0] is not None else ""
            count_value = _to_int(row[1]) if len(row) > 1 else None

            if label == "Category":
                in_category_table = True
                continue

            if not in_category_table:
                continue

            if label == "TOTAL":
                total_uses_formula = isinstance(row[1], str) and str(row[1]).strip().startswith("=")
                total_formula_text = str(row[1]).strip() if total_uses_formula else None
                observed_total = count_value if count_value is not None else sum(observed_counts.values())
                break

            if label in observed_counts:
                pytest.fail(f"Duplicate category row found in Summary Dashboard: {label}")

            if label in expected_counts:
                if count_value is None:
                    pytest.fail(
                        "Summary Dashboard category count must be numeric for "
                        f"{label}; found {row[1]!r}"
                    )
                observed_counts[label] = count_value
                continue

            if label:
                unknown_labels.append(label)
    finally:
        workbook.close()

    assert observed_counts == expected_counts, (
        "Summary Dashboard category totals do not match All Test Cases: "
        f"expected {expected_counts}, observed {observed_counts}"
    )
    assert not unknown_labels, f"Unexpected category label(s) in Summary Dashboard: {unknown_labels}"
    if total_uses_formula:
        assert total_formula_text is not None
        assert total_formula_text.replace(" ", "").upper() == "=SUM(B5:B10)", (
            "Summary Dashboard TOTAL formula changed unexpectedly: "
            f"observed {total_formula_text}"
        )
    assert total_uses_formula or observed_total is not None, "Summary Dashboard TOTAL row is missing a usable value"
    assert observed_total == expected_total, (
        "Summary Dashboard TOTAL does not match All Test Cases: "
        f"expected {expected_total}, observed {observed_total}"
    )


def test_scenario_ids_have_no_missing_sequence_gaps() -> None:
    """Prevent accidental scenario loss, e.g. DO-03 removed while DO-04+ still exist."""
    cases = _load_case_rows()
    ids = [str(case.get("Test ID") or "").strip() for case in cases if _non_empty(case.get("Test ID"))]

    numbers_by_prefix: dict[str, list[int]] = {}
    invalid_ids: list[str] = []

    for test_id in ids:
        match = SCENARIO_ID_PATTERN.match(test_id)
        if not match:
            invalid_ids.append(test_id)
            continue

        prefix, number = match.group(1), int(match.group(2))
        numbers_by_prefix.setdefault(prefix, []).append(number)

    assert not invalid_ids, f"Invalid scenario ID format(s): {sorted(invalid_ids)}"

    missing_by_prefix: dict[str, list[str]] = {}
    for prefix, numbers in numbers_by_prefix.items():
        existing = set(numbers)
        expected_range = set(range(1, max(numbers) + 1))
        missing = sorted(expected_range - existing)
        if missing:
            missing_by_prefix[prefix] = [f"{prefix}-{n:02d}" for n in missing]

    assert not missing_by_prefix, f"Missing scenario IDs detected: {missing_by_prefix}"


def test_conversation_script_scenarios_are_linked_to_main_case_list() -> None:
    script_case_ids = _load_conversation_script_case_ids()
    assert script_case_ids, "Conversation Scripts sheet should reference at least one scenario ID"

    main_case_ids = {
        str(case.get("Test ID") or "").strip()
        for case in _load_case_rows()
        if _non_empty(case.get("Test ID"))
    }

    missing_in_main = sorted(set(script_case_ids) - main_case_ids)
    assert not missing_in_main, (
        "Conversation Scripts references scenario IDs missing from All Test Cases: "
        f"{missing_in_main}"
    )


@pytest.mark.parametrize("test_id", sorted(EXPECTED_TEST_IDS))
def test_each_excel_scenario_is_well_formed(test_id: str) -> None:
    cases_by_id = {str(case.get("Test ID")).strip(): case for case in _load_case_rows() if _non_empty(case.get("Test ID"))}
    case = cases_by_id[test_id]

    for column in REQUIRED_COLUMNS:
        assert _non_empty(case.get(column)), f"{test_id} missing required column: {column}"

    category = str(case["Category"]).strip()
    priority = str(case["Priority"]).strip()

    assert "-" in test_id, f"{test_id} must include prefix-number format"
    prefix = test_id.split("-", 1)[0]
    assert prefix in EXPECTED_CATEGORY_BY_PREFIX, f"{test_id} has unsupported prefix"
    assert category == EXPECTED_CATEGORY_BY_PREFIX[prefix], (
        f"{test_id} category mismatch: expected {EXPECTED_CATEGORY_BY_PREFIX[prefix]}, got {category}"
    )
    assert priority in VALID_PRIORITIES, f"{test_id} has invalid priority: {priority}"
    assert len(str(case["Expected System Behaviour"]).strip()) >= 20, (
        f"{test_id} expected behaviour text is too short"
    )

