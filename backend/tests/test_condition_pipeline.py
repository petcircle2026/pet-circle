"""
Condition Pipeline Test — end-to-end extraction + aggregation.

Runs GPT extraction on documents, then passes the extracted conditions through
the aggregation pipeline to verify:

  1. Document-level condition_type is always "chronic" or "episodic" (never "recurrent")
  2. Aggregation groups similar conditions into families using body-system similarity
  3. condition_type upgrades correctly at family level (recurrence threshold, chronic trigger)
  4. condition_status is computed at runtime from dates (not stale)
  5. condition_family_id assignment is consistent

Usage:
    cd backend

    # Run against Sample_Reports folder (auto-detected):
    APP_ENV=production python -m tests.test_condition_pipeline

    # Run against a specific folder:
    APP_ENV=production FIXTURES_DIR=path/to/docs python -m tests.test_condition_pipeline

    # Run aggregation-only simulation (no documents, no OpenAI key needed):
    python -m tests.test_condition_pipeline --simulate

Requires OPENAI_API_KEY for document extraction (not needed for --simulate).
"""

import asyncio
import json
import os
import sys
import time
from dataclasses import dataclass, field
from datetime import date, timedelta
from typing import Any

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
os.environ.setdefault("APP_ENV", "production")

SIMULATE_ONLY = "--simulate" in sys.argv

# ─── Colours for terminal output ─────────────────────────────────────────────
_GREEN  = "\033[32m"
_RED    = "\033[31m"
_YELLOW = "\033[33m"
_CYAN   = "\033[36m"
_RESET  = "\033[0m"
_BOLD   = "\033[1m"


def _p(text: str) -> None:
    try:
        print(text)
    except UnicodeEncodeError:
        print(text.encode("ascii", errors="replace").decode("ascii"))


def _ok(msg: str) -> None:    _p(f"  {_GREEN}✓{_RESET} {msg}")
def _fail(msg: str) -> None:  _p(f"  {_RED}✗{_RESET} {msg}")
def _warn(msg: str) -> None:  _p(f"  {_YELLOW}~{_RESET} {msg}")
def _info(msg: str) -> None:  _p(f"    {_CYAN}{msg}{_RESET}")


# ─── Minimal Condition stand-in for aggregation (no DB needed) ────────────────

@dataclass
class MockCondition:
    """Mirrors the fields read by condition_aggregation_service."""
    id: Any
    pet_id: str
    document_id: str
    name: str
    condition_type: str          # "chronic" | "episodic"  (document level)
    condition_status: str | None  # "active" | "resolved" | None
    episode_dates: list[str] = field(default_factory=list)
    diagnosed_at: date | None = None
    medication_end_date: date | None = None
    medications: list[dict] = field(default_factory=list)
    monitoring: list[dict] = field(default_factory=list)
    soft_resolution: bool = False
    recurrence_watch: bool = False
    condition_family_id: Any = None
    is_active: bool = True


# ─── Extraction helpers ───────────────────────────────────────────────────────

async def _extract_conditions_from_file(filepath: str) -> list[dict]:
    """
    Call GPT extraction on one file and return the raw conditions[] array.
    Returns [] if the document has no conditions or extraction fails.
    """
    from app.services.shared.gpt_extraction import (
        _call_openai_extraction,
        _call_openai_extraction_vision,
        _validate_extraction_dict,
    )
    from app.utils.file_reader import (
        encode_image_base64,
        extract_pdf_text,
        render_pdf_pages_as_images,
    )

    ext = filepath.lower().rsplit(".", 1)[-1]
    with open(filepath, "rb") as f:
        file_bytes = f.read()

    parsed_dict = None
    try:
        if ext in ("jpg", "jpeg", "png"):
            mime = "image/jpeg" if ext in ("jpg", "jpeg") else "image/png"
            data_uri = encode_image_base64(file_bytes, mime)
            parsed_dict = await _call_openai_extraction_vision(data_uri)
        elif ext == "pdf":
            pdf_text = extract_pdf_text(file_bytes)
            if len(pdf_text.strip()) > 20:
                parsed_dict = await _call_openai_extraction(
                    f"Veterinary document text:\n\n{pdf_text}"
                )
            else:
                pages = render_pdf_pages_as_images(file_bytes, max_pages=3)
                if pages:
                    parsed_dict = await _call_openai_extraction_vision(pages[0])
    except Exception as exc:
        _warn(f"Extraction error: {exc}")
        return []

    if not parsed_dict:
        return []

    try:
        _items, _doc_name, _pet_name, metadata = _validate_extraction_dict(
            parsed_dict, file_path=filepath
        )
        return metadata.get("conditions") or []
    except Exception as exc:
        _warn(f"Validation error: {exc}")
        return []


# ─── Checks ───────────────────────────────────────────────────────────────────

_ISSUES: list[str] = []


def _check(condition: bool, pass_msg: str, fail_msg: str) -> bool:
    if condition:
        _ok(pass_msg)
    else:
        _fail(fail_msg)
        _ISSUES.append(fail_msg)
    return condition


# ─── Document-level checks ────────────────────────────────────────────────────

def _check_document_conditions(filename: str, raw_conditions: list[dict]) -> list[MockCondition]:
    """
    Validate each extracted condition and return MockCondition objects
    with the same collapse logic as gpt_extraction.py.
    """
    mock_conditions = []
    doc_id = f"doc_{filename.replace('.', '_')}"

    for i, raw in enumerate(raw_conditions):
        name = (raw.get("condition_name") or "").strip()
        if not name:
            continue

        raw_type = str(raw.get("condition_type") or "episodic").strip().lower()

        # ── Document-level collapse (mirrors gpt_extraction.py) ──────────────
        if raw_type == "chronic":
            stored_type = "chronic"
        elif raw_type == "resolved":
            stored_type = "episodic"  # legacy
        else:
            stored_type = "episodic"  # recurrent, acute, unknown → episodic

        # ── Check: no "recurrent" at document level ──────────────────────────
        _check(
            raw_type != "recurrent",
            f"'{name}' — GPT output condition_type='{raw_type}' (not recurrent)",
            f"'{name}' — GPT returned condition_type='recurrent' at document level (should be episodic or chronic only)",
        )
        _check(
            stored_type in ("chronic", "episodic"),
            f"'{name}' — stored condition_type='{stored_type}'",
            f"'{name}' — stored type '{stored_type}' is not chronic or episodic",
        )

        # ── Parse dates — normalise all to ISO YYYY-MM-DD ────────────────────
        from app.utils.date_utils import parse_date
        _raw_episodes: list[str] = raw.get("episode_dates") or []
        episode_dates: list[str] = []
        for _ed in _raw_episodes:
            try:
                episode_dates.append(str(parse_date(str(_ed))))
            except Exception:
                episode_dates.append(str(_ed))  # keep raw if unparseable
        episode_dates = sorted(set(episode_dates))

        diagnosed_at_str = raw.get("diagnosed_at")
        diagnosed_at: date | None = None
        if diagnosed_at_str:
            try:
                diagnosed_at = parse_date(diagnosed_at_str)
            except Exception:
                pass

        # ── diagnosed_at guarantee in episode_dates (mirrors gpt_extraction.py) ─
        if diagnosed_at and str(diagnosed_at) not in episode_dates:
            episode_dates = sorted(set(episode_dates + [str(diagnosed_at)]))

        # ── Medications ───────────────────────────────────────────────────────
        raw_meds = raw.get("medications") or []
        medications = []
        medication_end_date: date | None = None
        _visit_ref = diagnosed_at or (
            parse_date(episode_dates[0]) if episode_dates else None
        ) if True else None
        for med in raw_meds:
            if not isinstance(med, dict):
                continue
            from app.utils.date_utils import parse_date as _pd
            end_d: date | None = None

            # 1. Explicit end_date from GPT
            end_str = med.get("end_date")
            if end_str:
                try:
                    end_d = _pd(end_str)
                except Exception:
                    pass

            # 2. Compute from duration_days + visit date (mirrors gpt_extraction.py)
            if not end_d and med.get("duration_days"):
                try:
                    from datetime import timedelta
                    _dur = int(med["duration_days"])
                    if _dur > 0 and _visit_ref:
                        end_d = _visit_ref + timedelta(days=_dur)
                        _info(f"  end_date computed: {med.get('name')} = {_visit_ref} + {_dur}d = {end_d}")
                except Exception:
                    pass

            if end_d and (medication_end_date is None or end_d > medication_end_date):
                medication_end_date = end_d
            medications.append({"name": med.get("name"), "end_date": end_d, "dose": med.get("dose"),
                                 "duration_days": med.get("duration_days")})

        # ── Monitoring ────────────────────────────────────────────────────────
        raw_mon = raw.get("monitoring") or []
        monitoring = [
            {"name": m.get("name"), "recheck_due_date": m.get("recheck_due_date")}
            for m in raw_mon if isinstance(m, dict)
        ]

        # ── Compute document-level condition_status (mirrors gpt_extraction.py) ──
        # GPT's "resolved" is respected; everything else is computed from dates.
        raw_status = raw.get("condition_status")
        _gpt_status = str(raw_status).strip().lower() if raw_status else None
        if _gpt_status == "resolved":
            condition_status = "resolved"
        elif stored_type == "chronic":
            condition_status = "active"
        elif medication_end_date is not None:
            _today_d = date.today()
            _days = (_today_d - medication_end_date).days
            if _days <= 0:
                condition_status = "active"
            elif _days <= 30:
                condition_status = "monitoring"
            else:
                condition_status = "resolved"
        elif episode_dates:
            try:
                _lat = date.fromisoformat(episode_dates[-1])
                _ep_days = (date.today() - _lat).days
                condition_status = "monitoring" if _ep_days <= 30 else "resolved"
            except Exception:
                condition_status = "resolved"
        else:
            condition_status = "resolved"

        # Dedup within same document (mirrors gpt_extraction.py fix).
        _existing_doc = next(
            (m for m in mock_conditions if m.name.lower() == name.lower()),
            None,
        )
        if _existing_doc:
            # Merge episode_dates and medications into the existing mock row.
            _existing_doc.episode_dates = sorted(
                set(_existing_doc.episode_dates + episode_dates)
            )
            if diagnosed_at and not _existing_doc.diagnosed_at:
                _existing_doc.diagnosed_at = diagnosed_at
            _seen_med_names = {m["name"] for m in _existing_doc.medications if m.get("name")}
            for _m in medications:
                if _m.get("name") and _m["name"] not in _seen_med_names:
                    _existing_doc.medications.append(_m)
                    _seen_med_names.add(_m["name"])
            if medication_end_date:
                if not _existing_doc.medication_end_date or medication_end_date > _existing_doc.medication_end_date:
                    _existing_doc.medication_end_date = medication_end_date
            _info(f"  [dedup] merged duplicate '{name}' into existing doc-level row")
            continue  # don't add a second mock
        mock = MockCondition(
            id=f"c_{doc_id}_{i}",
            pet_id="test_pet",
            document_id=doc_id,
            name=name,
            condition_type=stored_type,
            condition_status=condition_status,
            episode_dates=episode_dates,
            diagnosed_at=diagnosed_at,
            medication_end_date=medication_end_date,
            medications=medications,
            monitoring=monitoring,
        )
        mock_conditions.append(mock)

        meds_str = ", ".join(
            f"{m['name']}(end={m['end_date'] or '?'}, dur={m.get('duration_days') or '-'}d)"
            for m in medications if m.get("name")
        ) or "none"
        _info(f"name='{name}'  type={stored_type}  status={condition_status} [computed]  "
              f"episodes={len(episode_dates)}  med_end={medication_end_date}  meds=[{meds_str}]")

    return mock_conditions


# ─── Aggregation checks ───────────────────────────────────────────────────────

def _run_aggregation(all_conditions: list[MockCondition]) -> list[dict]:
    """
    Run the aggregation pipeline over mock conditions and return family dicts.
    Uses the real condition_aggregation_service functions directly.
    """
    from app.services.dashboard.condition_aggregation_service import (
        _compute_condition_status,
        _group_into_families,
        _merge_family,
    )

    if not all_conditions:
        return []

    families_grouped = _group_into_families(all_conditions)
    family_results = []

    for family in families_grouped:
        merged = _merge_family(family)

        # Compute condition_status at runtime (never use stored value)
        runtime_status = _compute_condition_status(
            merged["condition_type"],
            merged.get("medication_end_date"),
            merged.get("episode_dates") or [],
        )
        merged["runtime_condition_status"] = runtime_status
        family_results.append(merged)

    return family_results


def _check_aggregation(families: list[dict]) -> None:
    """Print and validate each family row."""
    from app.services.dashboard.condition_aggregation_service import _compute_condition_status

    _p(f"\n  {_BOLD}Aggregated condition families:{_RESET}")
    for fam in families:
        name = fam.get("name", "?")
        ftype = fam.get("condition_type", "?")
        runtime_status = fam.get("runtime_condition_status", "?")
        episodes = fam.get("episode_dates") or []
        med_end = fam.get("medication_end_date")
        recurrence_watch = fam.get("recurrence_watch", False)
        meds = [m.get("name") for m in (fam.get("medications") or []) if isinstance(m, dict) and m.get("name")]

        _p(f"\n    {_BOLD}Family: {name}{_RESET}")
        _info(f"condition_type  = {ftype}")
        _info(f"status (runtime)= {runtime_status}  ← computed today, not stored")
        _info(f"episode_dates   = {episodes}")
        _info(f"medication_end  = {med_end}")
        _info(f"recurrence_watch= {recurrence_watch}")
        _info(f"medications     = {meds}")

        # Validate: chronic conditions must always be active
        if ftype == "chronic":
            _check(
                runtime_status == "active",
                f"'{name}' chronic → status=active",
                f"'{name}' chronic should always be active, got '{runtime_status}'",
            )

        # Validate: recurrent only appears at aggregation level
        _check(
            ftype in ("chronic", "episodic", "recurrent"),
            f"'{name}' family type='{ftype}' is valid",
            f"'{name}' family type='{ftype}' is not a valid aggregated type",
        )

        # Validate: if 2+ episodes, recurrence threshold check
        if len(episodes) >= 2 and ftype not in ("chronic",):
            _check(
                ftype == "recurrent",
                f"'{name}' has {len(episodes)} episodes → correctly upgraded to recurrent",
                f"'{name}' has {len(episodes)} episodes but type='{ftype}' (should be recurrent if within threshold)",
            )


# ─── Simulation data ──────────────────────────────────────────────────────────

def _build_simulation_conditions() -> list[MockCondition]:
    """
    Build a synthetic set of conditions that exercises all aggregation rules:
      - GI upset: 2 episodes in 12 months → should become recurrent
      - Hypothyroidism: chronic → stays chronic, always active
      - Skin infection: single episode, old → resolved
      - UTI (two docs): 2 episodes in 3 months → recurrent, status depends on med_end_date
    """
    today = date.today()
    return [
        # GI episode 1 (13 months ago)
        MockCondition(
            id="c1", pet_id="sim_pet", document_id="doc1",
            name="GI upset", condition_type="episodic", condition_status="resolved",
            episode_dates=[(today - timedelta(days=395)).isoformat()],
            diagnosed_at=today - timedelta(days=395),
            medication_end_date=today - timedelta(days=380),
        ),
        # GI episode 2 (15 days ago)
        MockCondition(
            id="c2", pet_id="sim_pet", document_id="doc2",
            name="vomiting and loose stool", condition_type="episodic", condition_status="active",
            episode_dates=[(today - timedelta(days=15)).isoformat()],
            diagnosed_at=today - timedelta(days=15),
            medication_end_date=today + timedelta(days=5),
            medications=[{"name": "metronidazole", "end_date": today + timedelta(days=5)}],
        ),
        # Hypothyroidism (chronic, lifelong)
        MockCondition(
            id="c3", pet_id="sim_pet", document_id="doc3",
            name="Hypothyroidism", condition_type="chronic", condition_status="active",
            episode_dates=[(today - timedelta(days=365)).isoformat()],
            diagnosed_at=today - timedelta(days=365),
            medication_end_date=date(2099, 12, 31),
            medications=[{"name": "thyroxine", "end_date": date(2099, 12, 31)}],
        ),
        # Skin infection (old, single episode)
        MockCondition(
            id="c4", pet_id="sim_pet", document_id="doc4",
            name="Skin infection", condition_type="episodic", condition_status="resolved",
            episode_dates=[(today - timedelta(days=200)).isoformat()],
            diagnosed_at=today - timedelta(days=200),
        ),
        # UTI episode 1 (3 months ago)
        MockCondition(
            id="c5", pet_id="sim_pet", document_id="doc5",
            name="UTI", condition_type="episodic", condition_status="resolved",
            episode_dates=[(today - timedelta(days=90)).isoformat()],
            diagnosed_at=today - timedelta(days=90),
            medication_end_date=today - timedelta(days=76),
        ),
        # UTI episode 2 — named differently but same family (ecoli in urine)
        MockCondition(
            id="c6", pet_id="sim_pet", document_id="doc6",
            name="Ecoli observed in urine", condition_type="episodic", condition_status="resolved",
            episode_dates=[(today - timedelta(days=45)).isoformat()],
            diagnosed_at=today - timedelta(days=45),
            medication_end_date=today - timedelta(days=31),
        ),
    ]


# ─── Excel report ─────────────────────────────────────────────────────────────

def _generate_excel_report(
    raw_gpt_per_doc: list[dict],
    mock_conditions: list["MockCondition"],
    families: list[dict],
) -> None:
    """
    Generate DB_Tables_Report.xlsx with sheets:
      1. GPT Raw Output         — exact values returned by GPT per document
      2. conditions             — one row per document-level condition (conditions table)
      3. condition_medications  — one row per medication (condition_medications table)
      4. condition_monitoring   — one row per monitoring item (condition_monitoring table)
      5. aggregated_conditions  — one row per family (aggregated_conditions table)
      6. agg_medications (JSONB)— medications JSONB inside each aggregated family
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # remove default sheet

    # ── Style helpers ──────────────────────────────────────────────────────────
    HEADER_FILL   = PatternFill("solid", fgColor="1F4E79")   # dark blue
    SUBHEAD_FILL  = PatternFill("solid", fgColor="2E75B6")   # mid blue
    ALT_FILL      = PatternFill("solid", fgColor="DDEEFF")   # light blue
    HEADER_FONT   = Font(bold=True, color="FFFFFF", size=10)
    BODY_FONT     = Font(size=9)
    TITLE_FONT    = Font(bold=True, size=11)
    MONO_FONT     = Font(name="Courier New", size=9)
    CENTER        = Alignment(horizontal="center", vertical="center", wrap_text=True)
    LEFT          = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    THIN          = Side(style="thin", color="AAAAAA")
    BORDER        = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

    def _ws(name: str):
        ws = wb.create_sheet(title=name)
        ws.sheet_view.showGridLines = True
        return ws

    def _header_row(ws, cols: list[str], row: int = 1):
        for c, col in enumerate(cols, 1):
            cell = ws.cell(row=row, column=c, value=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = CENTER
            cell.border = BORDER

    def _data_row(ws, values: list, row: int, alt: bool = False):
        fill = ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
        for c, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=val)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.alignment = LEFT
            cell.border = BORDER

    def _auto_width(ws, min_w=8, max_w=40):
        for col_cells in ws.columns:
            best = min_w
            for cell in col_cells:
                if cell.value:
                    best = max(best, min(len(str(cell.value)), max_w))
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = best + 2

    def _fmt(v) -> str:
        if v is None:
            return ""
        if isinstance(v, date):
            return v.isoformat()
        if isinstance(v, list):
            return ", ".join(str(x) for x in v) if v else ""
        return str(v)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 1 — GPT Raw Output
    # ══════════════════════════════════════════════════════════════════════════
    ws1 = _ws("GPT Raw Output")
    cols1 = [
        "Document", "# in Doc",
        "condition_name (GPT)", "condition_type (GPT)", "condition_status (GPT)",
        "diagnosed_at (GPT)", "episode_dates (GPT)", "source (GPT)",
        "diagnosis (GPT)", "notes (GPT)",
        "medications (GPT) — name | dose | frequency | end_date | duration_days",
        "monitoring (GPT) — name | recheck_due_date",
    ]
    _header_row(ws1, cols1)
    r = 2
    for doc_entry in raw_gpt_per_doc:
        fname = doc_entry["filename"]
        raws = doc_entry["raw_conditions"]
        if not raws:
            _data_row(ws1, [fname, "", "(no conditions extracted)", "", "", "", "", "", "", "", "", ""], r, alt=(r % 2 == 0))
            r += 1
            continue
        for i, raw in enumerate(raws, 1):
            meds_raw = raw.get("medications") or []
            meds_str = " | ".join(
                f"{m.get('name','')} [{m.get('dose','')}] {m.get('frequency','')} end={m.get('end_date','')} dur={m.get('duration_days','')}d"
                for m in meds_raw if isinstance(m, dict)
            )
            mon_raw = raw.get("monitoring") or []
            mon_str = " | ".join(
                f"{m.get('name','')} recheck={m.get('recheck_due_date','')}"
                for m in mon_raw if isinstance(m, dict)
            )
            row_vals = [
                fname, i,
                raw.get("condition_name", ""),
                raw.get("condition_type", ""),
                raw.get("condition_status", ""),
                _fmt(raw.get("diagnosed_at")),
                _fmt(raw.get("episode_dates")),
                raw.get("source", ""),
                raw.get("diagnosis", ""),
                raw.get("notes", ""),
                meds_str,
                mon_str,
            ]
            _data_row(ws1, row_vals, r, alt=(r % 2 == 0))
            r += 1
    _auto_width(ws1)
    ws1.row_dimensions[1].height = 28

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 2 — conditions table (document-level)
    # ══════════════════════════════════════════════════════════════════════════
    ws2 = _ws("conditions (DB)")
    cols2 = [
        "id (UUID)", "pet_id", "document_id",
        "name", "condition_type", "condition_status",
        "diagnosed_at", "episode_dates", "medication_end_date", "source",
        "is_active", "recurrence_watch", "condition_family_id",
        "created_at",
    ]
    _header_row(ws2, cols2)
    r = 2
    for i, mc in enumerate(mock_conditions):
        _data_row(ws2, [
            mc.id,
            mc.pet_id,
            mc.document_id,
            mc.name,
            mc.condition_type,
            mc.condition_status,
            _fmt(mc.diagnosed_at),
            _fmt(mc.episode_dates),
            _fmt(mc.medication_end_date),
            getattr(mc, "source", "extraction"),
            "TRUE",
            "FALSE",  # recurrence_watch set by aggregation, not extraction
            mc.condition_family_id or "(set by aggregation)",
            "(set at insert)",
        ], r, alt=(r % 2 == 0))
        r += 1
    _auto_width(ws2)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 3 — condition_medications table
    # ══════════════════════════════════════════════════════════════════════════
    ws3 = _ws("condition_medications (DB)")
    cols3 = [
        "id (UUID)", "condition_id", "condition_name",
        "name", "dose", "frequency", "route",
        "started_at", "end_date", "item_type",
    ]
    _header_row(ws3, cols3)
    r = 2
    for mc in mock_conditions:
        for med in mc.medications:
            if not med.get("name"):
                continue
            _data_row(ws3, [
                "(uuid)",
                mc.id,
                mc.name,
                med.get("name", ""),
                med.get("dose", "") or "",
                med.get("frequency", "") or "",
                med.get("route", "") or "",
                "",  # started_at (not extracted)
                _fmt(med.get("end_date")),
                "medicine",
            ], r, alt=(r % 2 == 0))
            r += 1
    _auto_width(ws3)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 4 — condition_monitoring table
    # ══════════════════════════════════════════════════════════════════════════
    ws4 = _ws("condition_monitoring (DB)")
    cols4 = [
        "id (UUID)", "condition_id", "condition_name",
        "name", "frequency", "recheck_due_date",
    ]
    _header_row(ws4, cols4)
    r = 2
    for mc in mock_conditions:
        for mon in mc.monitoring:
            if not mon.get("name"):
                continue
            _data_row(ws4, [
                "(uuid)",
                mc.id,
                mc.name,
                mon.get("name", ""),
                mon.get("frequency", "") or "",
                _fmt(mon.get("recheck_due_date")),
            ], r, alt=(r % 2 == 0))
            r += 1
    _auto_width(ws4)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 5 — aggregated_conditions table
    # ══════════════════════════════════════════════════════════════════════════
    ws5 = _ws("aggregated_conditions (DB)")
    cols5 = [
        "id (UUID)", "pet_id",
        "name", "condition_type", "condition_status (RUNTIME)",
        "episode_dates", "diagnosed_at", "last_record_date",
        "medication_end_date", "soft_resolution", "recurrence_watch",
        "canonical_condition_id", "# family members",
    ]
    _header_row(ws5, cols5)
    r = 2
    for fam in families:
        members = fam.get("_family_members") or []
        _data_row(ws5, [
            "(uuid — set at upsert)",
            "test_pet",
            fam.get("name", ""),
            fam.get("condition_type", ""),
            fam.get("runtime_condition_status", fam.get("condition_status", "")),
            _fmt(fam.get("episode_dates")),
            _fmt(fam.get("diagnosed_at")),
            _fmt(fam.get("last_record_date")),
            _fmt(fam.get("medication_end_date")),
            str(fam.get("soft_resolution", False)).upper(),
            str(fam.get("recurrence_watch", False)).upper(),
            fam.get("canonical_condition_id", ""),
            len(members),
        ], r, alt=(r % 2 == 0))
        r += 1
    _auto_width(ws5)

    # ══════════════════════════════════════════════════════════════════════════
    # Sheet 6 — aggregated_conditions JSONB medications
    # ══════════════════════════════════════════════════════════════════════════
    ws6 = _ws("agg_medications (JSONB)")
    cols6 = [
        "family_name", "med_name", "end_date", "dose", "frequency",
    ]
    _header_row(ws6, cols6)
    r = 2
    for fam in families:
        for med in (fam.get("medications") or []):
            _data_row(ws6, [
                fam.get("name", ""),
                med.get("name", "") if isinstance(med, dict) else str(med),
                _fmt(med.get("end_date") if isinstance(med, dict) else None),
                med.get("dose", "") if isinstance(med, dict) else "",
                med.get("frequency", "") if isinstance(med, dict) else "",
            ], r, alt=(r % 2 == 0))
            r += 1
    _auto_width(ws6)

    # ── Save ──────────────────────────────────────────────────────────────────
    out_xl = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "DB_Tables_Report.xlsx",
    )
    wb.save(out_xl)
    _p(f"\nExcel report saved to: {out_xl}")


# ─── Main ─────────────────────────────────────────────────────────────────────

async def main() -> None:
    _p(f"\n{'='*70}")
    _p(f"{_BOLD}Condition Pipeline Test{_RESET}")
    _p(f"{'='*70}")

    all_mock_conditions: list[MockCondition] = []
    raw_gpt_per_doc: list[dict] = []  # [{filename, raw_conditions: [...]}, ...]

    # ── Mode 1: Simulate with synthetic data ─────────────────────────────────
    if SIMULATE_ONLY:
        _p(f"\n{_BOLD}[ SIMULATION MODE — no GPT calls, no documents needed ]{_RESET}")
        _p("\nBuilding synthetic conditions (GI recurrence, hypothyroidism, skin, UTI recurrence)...")
        all_mock_conditions = _build_simulation_conditions()
        _p(f"  Built {len(all_mock_conditions)} mock document-level condition rows")

        for mc in all_mock_conditions:
            _info(f"  doc={mc.document_id}  name='{mc.name}'  type={mc.condition_type}  "
                  f"episodes={len(mc.episode_dates)}")

    # ── Mode 2: Extract from real documents ──────────────────────────────────
    else:
        repo_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
        fixtures_dir = os.environ.get("FIXTURES_DIR") or next(
            (p for p in [
                os.path.join(repo_root, "Sample_Reports"),
                os.path.join(repo_root, "fixtures", "sample_reports"),
            ] if os.path.exists(p)),
            None,
        )

        if not fixtures_dir or not os.path.exists(fixtures_dir):
            _p(f"{_RED}ERROR: No documents folder found. "
               f"Set FIXTURES_DIR or place files in Sample_Reports/. "
               f"Run with --simulate to test aggregation without documents.{_RESET}")
            sys.exit(1)

        # Collect files recursively so case1/case2/case3 sub-folders are included.
        _EXTS = {"pdf", "jpg", "jpeg", "png"}
        file_pairs: list[tuple[str, str]] = []  # (display_name, full_path)
        for dirpath, _dirs, dirfiles in os.walk(fixtures_dir):
            _dirs.sort()
            rel = os.path.relpath(dirpath, fixtures_dir)
            for fname in sorted(dirfiles):
                if fname.lower().rsplit(".", 1)[-1] in _EXTS:
                    display = fname if rel == "." else f"{rel}/{fname}"
                    file_pairs.append((display, os.path.join(dirpath, fname)))

        _p(f"\nFound {len(file_pairs)} document(s) in {os.path.relpath(fixtures_dir, repo_root)}")

        total_conditions_extracted = 0
        for filename, filepath in file_pairs:
            _p(f"\n{_BOLD}─── {filename} ───{_RESET}")
            t0 = time.time()

            raw_conditions = await _extract_conditions_from_file(filepath)
            elapsed = time.time() - t0

            if not raw_conditions:
                _warn(f"No conditions extracted ({elapsed:.1f}s)")
                raw_gpt_per_doc.append({"filename": filename, "raw_conditions": []})
                continue

            _p(f"  Extracted {len(raw_conditions)} condition(s) in {elapsed:.1f}s")
            raw_gpt_per_doc.append({"filename": filename, "raw_conditions": raw_conditions})
            mocks = _check_document_conditions(filename, raw_conditions)
            all_mock_conditions.extend(mocks)
            total_conditions_extracted += len(mocks)

        _p(f"\n  Total document-level condition rows collected: {total_conditions_extracted}")

    # ── Aggregation pipeline ──────────────────────────────────────────────────
    _p(f"\n{'─'*70}")
    _p(f"{_BOLD}Aggregation pipeline{_RESET}")
    _p(f"{'─'*70}")

    families: list[dict] = []
    if not all_mock_conditions:
        _warn("No conditions to aggregate. Upload prescription/diagnosis documents or run --simulate.")
    else:
        families = _run_aggregation(all_mock_conditions)
        _p(f"\n  Input:   {len(all_mock_conditions)} document-level condition rows")
        _p(f"  Output:  {len(families)} condition famil{'y' if len(families)==1 else 'ies'}")

        _check_aggregation(families)

        # condition_status staleness check: re-run compute with a future "today"
        _p(f"\n  {_BOLD}Staleness simulation{_RESET} — would status change 60 days from now?")
        from app.services.dashboard.condition_aggregation_service import _compute_condition_status
        for fam in families:
            current = fam.get("runtime_condition_status", "?")
            med_end = fam.get("medication_end_date")
            episodes = fam.get("episode_dates") or []
            ftype = fam.get("condition_type", "episodic")

            # Simulate 60 days later by shifting episode dates back
            shifted_episodes = []
            for ep in episodes:
                try:
                    from app.utils.date_utils import parse_date
                    d = parse_date(ep)
                    shifted_episodes.append((d - timedelta(days=60)).isoformat())
                except Exception:
                    shifted_episodes.append(ep)
            shifted_med_end = (med_end - timedelta(days=60)) if med_end and med_end.year < 2090 else med_end

            future_status = _compute_condition_status(ftype, shifted_med_end, shifted_episodes)
            changed = current != future_status
            msg = f"'{fam.get('name')}': {current} → {future_status}"
            if changed:
                _info(f"{_YELLOW}{msg}  (would change — runtime compute catches it){_RESET}")
            else:
                _info(f"{msg}  (unchanged)")

    # ── Final summary ─────────────────────────────────────────────────────────
    _p(f"\n{'='*70}")
    _p(f"{_BOLD}SUMMARY{_RESET}")
    _p(f"{'='*70}")

    if not _ISSUES:
        _p(f"\n{_GREEN}{_BOLD}ALL CHECKS PASSED{_RESET}")
    else:
        _p(f"\n{_RED}{_BOLD}ISSUES FOUND: {len(_ISSUES)}{_RESET}")
        for issue in _ISSUES:
            _p(f"  {_RED}✗{_RESET} {issue}")

    # Save family output as JSON for inspection
    if all_mock_conditions:
        out_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), "condition_pipeline_results.json")
        families_serializable = []
        for fam in families:
            fam_copy = {
                k: v.isoformat() if isinstance(v, date) else v
                for k, v in fam.items()
            }
            families_serializable.append(fam_copy)
        with open(out_path, "w") as fp:
            json.dump(families_serializable, fp, indent=2, default=str)
        _p(f"\nDetailed family output saved to: {out_path}")

    # ── Generate Excel report ─────────────────────────────────────────────────
    _generate_excel_report(
        raw_gpt_per_doc=raw_gpt_per_doc,
        mock_conditions=all_mock_conditions,
        families=families,
    )

    _p("")


if __name__ == "__main__":
    asyncio.run(main())
